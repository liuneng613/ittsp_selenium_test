import openpyxl
from openpyxl import Workbook

from sePublic import CustomizationLog
from sePublic import ReadConfig

my_logger = CustomizationLog.CustomizationLog("LoginCase")


class ReadWriteExcelData:
    """
    读取和写入excel表数据
    """
    def __init__(self, path):
        # 初始化路径与Workbook对象
        self.path = path
        self.wb = Workbook()

    '''
    读取excel表数据，返回excel指定表的所有数据，以嵌套列表返回
    '''
    def read_data(self, sheet_name):
        # 打开EXCEL文件
        excel = openpyxl.load_workbook(self.path)
        # 获取指定名称的工作表
        sheet = excel.get_sheet_by_name(sheet_name)
        # print(type((sheet.cell(row=1, column=1)).value))
        # print(type(sheet.max_row))
        # print(type(sheet.max_column))
        # 创建一个空test_data
        test_data = []
        # 从第二行开始遍历数据存入row_list中
        for row in range(1, sheet.max_row + 1):
            row_list = []
            for col in range(1, sheet.max_column + 1):
                row_list.append(sheet.cell(row=row, column=col).value)
            test_data.append(row_list)
        return test_data
    '''
    excel表数据写入，传参：data：列表，sheet_name：表名
    '''
    def write_data(self, data, sheet_name):
        # 等同于  ws = wb.get_active_sheet()，定位到sheet
        sheet_function = self.wb.active
        # 设置sheet表名
        sheet_function.title = sheet_name
        # 按行写入数据
        for row in data:
            sheet_function.append(row)
        # 保存数据到指定目录
        self.wb.save(ReadConfig.root_path + r'\results\results_excel\%s.xlsx' % sheet_name)
    '''
    查找指定list数据data，第一次为col_name的索引index，（就是在excel中查找指定列名的列数）
    '''
    def col_index(self, data, col_name):
        try:
            for col in data[0]:
                if col == col_name:
                    return data[0].index(col)
        except TypeError as msg:
            my_logger.info('TypeError：' + msg)

if __name__ == '__main__':
    test_excel = ReadWriteExcelData(ReadConfig.root_path + r"\test_case\user_test.xlsx")
    print(ReadConfig.root_path + r"\test_case\user_test.xlsx")
    print(test_excel.read_data('login'))
    test_excel.write_data(test_excel.read_data('login'), 'login')
    print(test_excel.col_index(test_excel.read_data('login'), '预期结果'))